import yfinance as yf
import pandas as pd
import datetime
minha_lista = []
# Definir os ativos (incluindo AMX, AAPL e PETR4)
ativos = ["AMX", "AAPL", "PETR4.SA", "ITUB3.SA", "VALE3.SA", "PETR3.SA", "^BVSP"]

# Definir o período para hoje (tempo real)
hoje = datetime.datetime.now()
data_inicial = hoje - datetime.timedelta(days=1)  # Últimos 2 dias
data_final = hoje

try:
    # Baixar os dados
    dados = yf.download(ativos, start=data_inicial, end=data_final, group_by='ticker')
    
    if dados.empty:
        print("Nenhum dado encontrado. Verifique os tickers e sua conexão com a internet.")
    else:
        # Mostrar os dados mais recentes (último pregão)
        print("\nÚltimas cotações:")
        for acao in ativos:
            if acao in dados:
                ultimo_preco = dados[acao].iloc[-1]['Close']
                print(f"{acao}: {ultimo_preco:.2f}")
    
        # Se quiser ver todo o dataframe com todos os dados
        #print("\nTodos os dados baixados:")
        #print(dados)
        
except Exception as e:
    print(f"Erro ao buscar dados: {str(e)}")
    print("Possíveis causas:")
    print("- Algum ticker pode estar incorreto")
    print("- Problema de conexão com a internet")
    print("- Limitação da API (muitas requisições)")




import pandas as pd
# df = pd.DataFrame(data)
# arquivo_excel = "dados_pessoas.xlsx"
# df.to_excel(arquivo_excel, index=False)
#print(f"Dados salvos em '{arquivo_excel}'.")
#print(data)

import yfinance as yf

def get_ultimo_valor_acao(ticker):
  try:
    acao = yf.Ticker(ticker)
    historico = acao.history(period="1d")
    ultimo_valor = historico.iloc[-1]['Close']
    return ultimo_valor
  except Exception as e:
    print(f"Erro ao obter o valor da ação: {e}")
    return None

ticker_acao = "ITUB4.SA"
valor = get_ultimo_valor_acao(ticker_acao)

print(f"{ticker_acao} : {valor}")

minha_lista.append(valor)

print(minha_lista)


import openpyxl

def inserir_lista_no_excel(lista, caminho_arquivo, nome_aba="Plan1"):

    if nome_aba in workbook.sheetnames:
        aba = workbook[nome_aba]
    else:
        aba = workbook.create_sheet(title=nome_aba)

    # Itera sobre a lista e escreve os elementos nas células
    for i, elemento in enumerate(lista):
        aba.cell(row=i+1, column=1, value=elemento)  # Insere na coluna 1

    # Salva as alterações no arquivo
    workbook.save(caminho_arquivo)
    print(f"Lista inserida com sucesso no arquivo: {caminho_arquivo}")


# Exemplo de uso:
caminho = "valores.xlsx"
inserir_lista_no_excel(minha_lista, caminho)